"""Tests for Excel report generation."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from ils_kickoff.analyses import run_all_analyses
from ils_kickoff.charts import create_charts
from ils_kickoff.data_loader import load_data
from ils_kickoff.excel_report import write_excel_report
from ils_kickoff.pipeline import PipelineResult


@pytest.fixture
def pipeline_result(sample_settings):
    """Run the full pipeline for report testing."""
    df = load_data(sample_settings)
    analyses = run_all_analyses(df, sample_settings)
    charts = create_charts(analyses, sample_settings)
    sample_settings.output_dir.mkdir(parents=True, exist_ok=True)
    return PipelineResult(
        settings=sample_settings, df=df, analyses=analyses, charts=charts,
    )


class TestExcelReport:
    """Test Excel report generation."""

    def test_creates_file(self, pipeline_result, tmp_path):
        path = tmp_path / "test_report.xlsx"
        write_excel_report(pipeline_result, path)
        assert path.exists()
        assert path.stat().st_size > 0

    def test_has_report_info_sheet(self, pipeline_result, tmp_path):
        path = tmp_path / "test_report.xlsx"
        write_excel_report(pipeline_result, path)
        wb = load_workbook(path)
        assert "Report Info" in wb.sheetnames
        wb.close()

    def test_has_analysis_sheets(self, pipeline_result, tmp_path):
        path = tmp_path / "test_report.xlsx"
        write_excel_report(pipeline_result, path)
        wb = load_workbook(path)
        successful = [a for a in pipeline_result.analyses if a.error is None]
        # Should have Report Info + TOC + analysis sheets
        assert len(wb.sheetnames) >= len(successful) + 1
        wb.close()

    def test_frozen_panes(self, pipeline_result, tmp_path):
        path = tmp_path / "test_report.xlsx"
        write_excel_report(pipeline_result, path)
        wb = load_workbook(path)
        # Check that at least some analysis sheets have frozen panes
        skip_sheets = {"Report Info", "Table of Contents"}
        frozen_count = 0
        for ws in wb.worksheets:
            if ws.title not in skip_sheets and ws.max_row > 1:
                if ws.freeze_panes is not None:
                    frozen_count += 1
        assert frozen_count > 0, "No analysis sheets have frozen panes"
        wb.close()
