"""Formatted Excel report generation with NamedStyles."""

import logging
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from ils_kickoff.formatting import (
    excel_number_format,
    is_grand_total_row,
    is_percentage_column,
)

logger = logging.getLogger(__name__)

# Style constants
NAVY = "1B365D"
LIGHT_BLUE = "D6E4F0"
ZEBRA_GRAY = "FAFAFA"
TOTAL_GRAY = "F0F0F0"
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def _register_styles(wb: Workbook) -> None:
    """Register NamedStyles once for batch application."""
    header_style = NamedStyle(name="rpt_header")
    header_style.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_style.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_style.border = THIN_BORDER
    wb.add_named_style(header_style)

    even_style = NamedStyle(name="rpt_data_even")
    even_style.font = Font(name="Calibri", size=10)
    even_style.alignment = Alignment(horizontal="center", vertical="center")
    even_style.border = THIN_BORDER
    wb.add_named_style(even_style)

    odd_style = NamedStyle(name="rpt_data_odd")
    odd_style.font = Font(name="Calibri", size=10)
    odd_style.fill = PatternFill(start_color=ZEBRA_GRAY, end_color=ZEBRA_GRAY, fill_type="solid")
    odd_style.alignment = Alignment(horizontal="center", vertical="center")
    odd_style.border = THIN_BORDER
    wb.add_named_style(odd_style)

    total_style = NamedStyle(name="rpt_total")
    total_style.font = Font(name="Calibri", bold=True, size=10)
    total_style.fill = PatternFill(start_color=TOTAL_GRAY, end_color=TOTAL_GRAY, fill_type="solid")
    total_style.alignment = Alignment(horizontal="center", vertical="center")
    total_style.border = THIN_BORDER
    wb.add_named_style(total_style)


def _write_cover_sheet(wb: Workbook, result) -> None:
    """Write the Report Info cover sheet."""
    ws = wb.active
    ws.title = "Report Info"
    ws.sheet_properties.showGridLines = False

    # Title
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "ILS Kickoff Report"
    cell.font = Font(name="Calibri", size=24, bold=True, color=NAVY)

    ws.merge_cells("A2:D2")
    ws["A2"].value = result.settings.client_name
    ws["A2"].font = Font(name="Calibri", size=16, color="666666")

    # Report details
    now = datetime.now()
    details = [
        ("Client ID:", result.settings.client_id),
        ("Report Date:", now.strftime("%B %d, %Y")),
        ("Source File:", result.settings.data_file.name),
        ("Total Accounts:", f"{len(result.df):,}"),
        ("Analyses Run:", str(sum(1 for a in result.analyses if a.error is None))),
    ]

    for i, (label, value) in enumerate(details, start=4):
        ws[f"A{i}"].value = label
        ws[f"A{i}"].font = Font(name="Calibri", bold=True, size=11)
        ws[f"B{i}"].value = value
        ws[f"B{i}"].font = Font(name="Calibri", size=11)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50


def _write_toc_sheet(wb: Workbook, result) -> None:
    """Write Table of Contents with hyperlinks to each analysis sheet."""
    ws = wb.create_sheet("Contents", 1)
    ws.sheet_properties.showGridLines = False

    ws.merge_cells("A1:C1")
    ws["A1"].value = "Table of Contents"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color=NAVY)

    ws["A3"].value = "#"
    ws["B3"].value = "Analysis"
    ws["C3"].value = "Sheet"
    for cell in [ws["A3"], ws["B3"], ws["C3"]]:
        cell.font = Font(name="Calibri", bold=True, size=11)

    row = 4
    for i, analysis in enumerate(result.analyses, start=1):
        if analysis.error is not None:
            continue
        ws[f"A{row}"].value = i
        ws[f"B{row}"].value = analysis.title
        ws[f"C{row}"].value = analysis.sheet_name
        ws[f"C{row}"].hyperlink = f"#{analysis.sheet_name}!A1"
        ws[f"C{row}"].font = Font(name="Calibri", size=11, color="0563C1", underline="single")
        row += 1

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 25


def _write_analysis_sheet(
    wb: Workbook,
    analysis,
    chart_png: bytes | None = None,
) -> None:
    """Write a single analysis as a formatted worksheet."""
    df = analysis.df
    if df.empty:
        return

    ws = wb.create_sheet(analysis.sheet_name)
    ws.freeze_panes = "A2"

    # Header row
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.style = "rpt_header"

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        is_total = is_grand_total_row(row)
        is_odd = (row_idx % 2) == 1

        for col_idx, col_name in enumerate(df.columns, start=1):
            val = row[col_name]

            # Convert percentage display values for Excel formatting
            if is_percentage_column(col_name) and isinstance(val, (int, float)) and not pd.isna(val):
                val = val / 100.0  # Excel multiplies by 100 when displaying %

            cell = ws.cell(row=row_idx, column=col_idx, value=val)

            if is_total:
                cell.style = "rpt_total"
            elif is_odd:
                cell.style = "rpt_data_odd"
            else:
                cell.style = "rpt_data_even"

            # Apply number format
            cell.number_format = excel_number_format(col_name)

    # Autofilter
    if len(df) > 0:
        last_col = get_column_letter(len(df.columns))
        ws.auto_filter.ref = f"A1:{last_col}{len(df) + 1}"

    # Column widths (sample-based)
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = len(str(col_name))
        for val in df[col_name].head(20):
            display_len = len(str(val)) if not pd.isna(val) else 0
            # Account for number formatting adding chars
            if is_percentage_column(col_name):
                display_len += 2
            max_len = max(max_len, display_len)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

    # Embed chart image if available
    if chart_png:
        img = XLImage(BytesIO(chart_png))
        img.width = 700
        img.height = 400
        ws.add_image(img, f"A{len(df) + 4}")

    # Back to Contents link
    link_row = len(df) + 3 if not chart_png else len(df) + 25
    ws.cell(row=link_row, column=1, value="Back to Contents")
    ws.cell(row=link_row, column=1).hyperlink = "#Contents!A1"
    ws.cell(row=link_row, column=1).font = Font(
        name="Calibri", size=10, color="0563C1", underline="single"
    )


def write_excel_report(result, output_path: Path) -> None:
    """Write the complete Excel report."""
    wb = Workbook()
    _register_styles(wb)
    _write_cover_sheet(wb, result)
    _write_toc_sheet(wb, result)

    # Render chart PNGs
    chart_pngs: dict[str, bytes] = {}
    if result.charts:
        try:
            from ils_kickoff.charts import render_chart_png
            for name, fig in result.charts.items():
                try:
                    chart_pngs[name] = render_chart_png(fig, result.settings.charts)
                except Exception as e:
                    logger.warning("Chart PNG for '%s' failed: %s", name, e)
        except ImportError:
            pass

    # Write analysis sheets
    for analysis in result.analyses:
        if analysis.error is not None:
            continue
        _write_analysis_sheet(
            wb,
            analysis,
            chart_png=chart_pngs.get(analysis.name),
        )

    wb.save(output_path)
    logger.info("Excel report saved: %s", output_path)
